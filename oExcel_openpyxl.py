import openpyxl,os
exc_path=os.path.abspath('pythonLearn/other/testcases.xlsx') 
print(exc_path,type(exc_path)) 
# 读excel
data=openpyxl.load_workbook(exc_path)
print(type(data))  # <class 'openpyxl.workbook.workbook.Workbook'>
# 获取工作表
tables=data.get_sheet_names() # list形式
print(tables,type(tables)) # ['login', 'lines'] <class 'list'>
table=data.get_sheet_by_name('login') # 根据名称查询sheet
print(table,type(table)) # <Worksheet "login"> <class 'openpyxl.worksheet.worksheet.Worksheet'>
sheetname=table.title  # # 获取工作表的表名
print(sheetname)
# 获取行列
'''
openpyxl中便提供了一个行和列的生成器(sheet.rows和sheet.columns),这两个生成器里面是每一行（或列）的数据，每一行（或列）又由一个tuple包裹
'''
# 行
print(type(table.rows))# <class 'generator'>
for r in table.rows:
    print(r,type(r)) #  <class 'tuple'>
    for cell in r:
        print(cell.value)
l=[cell.value for cell in list(table.rows)[0]] # 返回第一行的值 list
#列 遍历每一列的值
for c in table.columns:
    for cell in c:
        print(cell.value)
c=[cell.value for cell in list(table.columns)[1]]# 返回第2列的值 list
print(c)
# #最大行和最大列
rows=table.max_row
cols=table.max_column
print(rows,cols)

# 写
table=data.get_sheet_by_name('test')
table['C1'].value='fail'
row = [1 ,2, 3, 4, 5]
table.append(row)
rows = [
    ['ID', 'Name', 'Department'],
    ['001', 'Lee','CS'],
    ['002', 'John','MA'],
    ['003', 'Amy','IS']
]
for row in rows:  # 只支持按行逐一写入，多行写入还需通过遍历的形式实现
    table.append(row)
data.save(exc_path)

class oExcel(object):
    def __init__(self):
        self.path=exc_path
        self.data=openpyxl.load_workbook(self.path)
    def rdExcel(self):
        sheets=self.data.get_sheet_names()
        sheetDict={}
        for sheet in sheets:
            table=self.data.get_sheet_by_name(sheet)
            rows=table.max_row
            if rows>1:
                key=[cell.value for cell in list(table.rows)[0]]
                sheetList=[]
                for i in range(1,rows):
                    row_value=[cell.value for cell in list(table.rows)[i]]
                    sheetList.append(dict(zip(key,row_value)))
                    i+=1
                sheetDict[sheet]=sheetList
            else:
                sheetDict[sheet]=[]
        return sheetDict
    def getData(self,sheetname):
        return self.rdExcel()[sheetname]


